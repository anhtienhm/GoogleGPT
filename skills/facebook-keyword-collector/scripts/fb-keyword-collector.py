#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FB Keyword Collector — thu thập thông tin Facebook theo từ khoá.

Điều khiển Chrome thật (profile đã đăng nhập FB) qua AppleScript + execute javascript.
Pipeline chính (source=google, mặc định):
  1. Tìm URL Facebook qua Google search trong Chrome:  site:facebook.com "<từ khoá>"
  2. Lọc URL (group / page / post / people)
  3. Mở từng URL trong 1 tab riêng, extract: tên, meta (thành viên/follower), nội dung, tác giả, thời gian
  4. (--comments) Mở bài viết, scroll + click "Xem thêm", thu thập bình luận + SĐT
  5. Xuất Excel (sheet "Kết quả") + JSON

Source=fb (thử nghiệm): search trực tiếp facebook.com/search/posts — FB 2026 hay trả decoy
(text bị chèn zero-width chars) khi phát hiện automation → script tự nhận và báo.

Usage:
  python3 fb-keyword-collector.py --keywords "sàn gỗ" "thi công sàn gỗ"
  python3 fb-keyword-collector.py --keywords "sàn gỗ" --type groups --limit 10
  python3 fb-keyword-collector.py --keywords "sàn gỗ" --comments --max-comment-posts 3
  python3 fb-keyword-collector.py --keywords "sàn gỗ" --source fb
  python3 fb-keyword-collector.py --check
"""
import argparse
import json
import re
import subprocess
import tempfile
import time
import urllib.parse
from datetime import datetime
from pathlib import Path

OUT_DIR = Path.home() / "Facebook"
DEFAULT_OUT = OUT_DIR / f"fb_keyword_{datetime.now():%Y%m%d_%H%M%S}.xlsx"

# ================= AppleScript helpers =================

def osa(script: str, timeout: int = 90) -> str:
    tmp = Path(tempfile.mktemp(suffix=".applescript"))
    tmp.write_text(script, encoding="utf-8")
    try:
        r = subprocess.run(["osascript", str(tmp)], capture_output=True, text=True, timeout=timeout)
    finally:
        try:
            tmp.unlink()
        except OSError:
            pass
    if r.returncode != 0:
        err = (r.stderr or "").strip()
        if "-1743" in err:
            raise RuntimeError(
                "macOS chặn quyền điều khiển Chrome. Vào System Settings → Privacy & Security → "
                "Automation, bật quyền điều khiển Google Chrome cho app đang chạy (Hermes/Terminal)."
            )
        if "not allowed" in err.lower() or "not authorized" in err.lower():
            raise RuntimeError(
                'Chrome chưa bật "Allow JavaScript from Apple Events": '
                "menu Chrome → View → Developer → Allow JavaScript from Apple Events.\n" + err
            )
        raise RuntimeError(f"osascript thất bại: {err}")
    return r.stdout


def frontmost_app() -> str:
    try:
        return osa(
            'tell application "System Events"\n'
            "    set f to name of first application process whose frontmost is true\n"
            "    return f\n"
            "end tell",
            timeout=30,
        ).strip()
    except RuntimeError:
        return ""


def ensure_chrome() -> None:
    # tell application "Google Chrome" tự launch Chrome nếu chưa chạy
    try:
        osa('tell application "Google Chrome" to count windows', timeout=30)
    except RuntimeError:
        subprocess.run(["open", "-a", "Google Chrome"], timeout=30)
        time.sleep(5)


def raise_chrome() -> None:
    """Kéo window 1 của Chrome lên trước — cần thiết để FB render lazy content
    (visibilityState phải = visible thì IntersectionObserver mới chạy)."""
    osa(
        'tell application "Google Chrome"\n'
        "    if (count of windows) > 0 then\n"
        "        set index of front window to 1\n"
        "    end if\n"
        "end tell\n"
        'tell application "Google Chrome" to activate',
        timeout=30,
    )


def ensure_tab() -> str:
    """Tạo tab riêng (about:blank), trả về tab id (bền hơn index)."""
    out = osa(
        'tell application "Google Chrome"\n'
        "    tell front window\n"
        '        make new tab with properties {URL:"about:blank"}\n'
        "        return (id of last tab)\n"
        "    end tell\n"
        "end tell",
        timeout=30,
    )
    return out.strip()


def activate_tab(tabid: str) -> None:
    """Activate tab theo id (so sánh text — so sánh số bị lỗi -10006)."""
    osa(
        f'tell application "Google Chrome"\n'
        "    tell front window\n"
        "        set i to 0\n"
        "        repeat with t in tabs\n"
        "            set i to i + 1\n"
        f'            if ((id of t) as text) = "{tabid}" then\n'
        "                set active tab index to i\n"
                "                exit repeat\n"
        "            end if\n"
        "        end repeat\n"
        "    end tell\n"
        "end tell",
        timeout=30,
    )


def nav(url: str, tabid: str) -> None:
    try:
        osa(
            f'tell application "Google Chrome" to set URL of tab id {tabid} of front window to "{url}"',
            timeout=60,
        )
    except RuntimeError:
        # tab bị đóng → tạo lại
        raise


def js(code: str, tabid: str) -> str:
    activate_tab(tabid)
    esc = code.replace("\\", "\\\\").replace('"', '\\"')
    script = (
        'tell application "Google Chrome"\n'
        "    tell active tab of front window\n"
        f'        set r to execute javascript "{esc}"\n'
        "    end tell\n"
        "end tell\n"
        "return r"
    )
    return osa(script)


def wait_ready(tabid: str, extra: float = 4.0, tries: int = 15) -> None:
    for _ in range(tries):
        try:
            if js("document.readyState", tabid).strip() == "complete":
                break
        except RuntimeError:
            pass
        time.sleep(1)
    time.sleep(extra)


# ================= JS snippets =================

LOGIN_CHECK_JS = r"""(function(){
  if (document.querySelector('input[name="email"], input[name="pass"], #loginbutton')) {
    return 'LOGIN_REQUIRED';
  }
  return 'OK';
})()"""

# Trang tổng quát: title, meta (thành viên/follower), text head, broken detect
PAGE_JS = r"""(function(){
  var out = {title:'', name:'', meta:'', text:'', kind:'', broken:false};
  var body = (document.body ? document.body.innerText : '') || '';
  body = body.replace(/\s+/g, ' ');
  out.title = document.title || '';
  if (body.indexOf('Trang này không hiển thị') !== -1) out.broken = true;
  var t = out.title.replace(/^\(\d+\)\s*/, '').replace(/\s*\|\s*Facebook\s*$/, '').trim();
  out.name = t;
  var m = body.match(/([\d.,]+[KkMm]?\s*(?:thành viên|thành viên đã tham gia|người theo dõi|lượt thích))/);
  if (m) out.meta = m[1];
  if (body.indexOf('Nhóm Công khai') !== -1) out.kind = 'group-public';
  else if (body.indexOf('Nhóm kín') !== -1) out.kind = 'group-closed';
  else if (body.indexOf('Nhóm riêng tư') !== -1) out.kind = 'group-private';
  else if (body.indexOf('Trang ·') !== -1 || body.indexOf('Sản phẩm/dịch vụ') !== -1 || body.indexOf('Doanh nghiệp địa phương') !== -1) out.kind = 'page';
  out.text = body.slice(0, 1500);
  return JSON.stringify(out);
})()"""

# Article chính của trang post: text bài + tác giả + thời gian
POST_PAGE_JS = r"""(function(){
  var out = {postText:'', author:'', time:''};
  var arts = document.querySelectorAll('div[role="article"]');
  var main = arts[0];
  if (!main) return JSON.stringify(out);
  var msg = main.querySelector('[data-ad-preview="message"]');
  out.postText = (msg ? msg.innerText : main.innerText || '').replace(/\s+/g, ' ').trim().slice(0, 2000);
  var links = main.querySelectorAll('a[href]');
  for (var i = 0; i < links.length; i++) {
    var h = links[i].href || '';
    var path = h.split('?')[0];
    var lt = (links[i].innerText || '').trim().replace(/\s+/g, ' ');
    if (!out.author && lt.length > 0 && lt.length < 50 && path.indexOf('facebook.com') !== -1 &&
        path.indexOf('/search') === -1 && path.indexOf('/posts/') === -1 && path.indexOf('/hashtag/') === -1 &&
        path.indexOf('/groups/') === -1 && path.indexOf('/videos/') === -1) out.author = lt;
    if (!out.author && lt.length > 0 && lt.length < 50 && path.indexOf('/user/') !== -1) out.author = lt;
    if (!out.time) {
      var al = links[i].getAttribute('aria-label') || '';
      if (al.length < 60 && /giờ|phút|ngày|tháng|năm|lúc|hôm qua/i.test(al)) out.time = al;
    }
  }
  return JSON.stringify(out);
})()"""

# Bình luận: bỏ article 0 (bài viết gốc), lấy author + text
COMMENTS_JS = r"""(function(){
  var out = [];
  var seen = {};
  var arts = document.querySelectorAll('div[role="article"]');
  for (var i = 1; i < arts.length; i++) {
    var art = arts[i];
    var t = (art.innerText || '').replace(/\s+/g, ' ').trim();
    if (t.length < 5) continue;
    var author = '';
    var authorUrl = '';
    var links = art.querySelectorAll('a[href]');
    for (var j = 0; j < links.length; j++) {
      var h = (links[j].href || '').split('?')[0];
      var lt = (links[j].innerText || '').trim().replace(/\s+/g, ' ');
      if (!author && lt.length > 0 && lt.length < 50 && h.indexOf('facebook.com') !== -1 &&
          h.indexOf('/search') === -1 && h.indexOf('/groups/') === -1 && h.indexOf('/posts/') === -1 &&
          h.indexOf('/hashtag/') === -1 && h.indexOf('/videos/') === -1) { author = lt; authorUrl = h; }
      if (!author && lt.length > 0 && lt.length < 50 && h.indexOf('/user/') !== -1) { author = lt; authorUrl = h; }
    }
    if (!author) continue;
    var key = author + '|' + t.slice(0, 40);
    if (seen[key]) continue;
    seen[key] = 1;
    out.push({author: author, authorUrl: authorUrl, text: t.slice(0, 1000)});
  }
  return JSON.stringify(out);
})()"""

# Click các nút "Xem thêm" (expand text + comments)
EXPAND_JS = r"""(function(){
  var n = 0;
  var btns = document.querySelectorAll('div[role="button"], span[role="button"]');
  for (var i = 0; i < btns.length; i++) {
    var t = (btns[i].innerText || '').trim();
    if (t.length < 40 && /xem thêm|see more|hiển thị thêm/i.test(t)) {
      try { btns[i].click(); n++; } catch (e) {}
    }
  }
  return n;
})()"""

# Lấy URL FB từ trang Google search đang mở
GOOGLE_LINKS_JS = r"""(function(){
  var out = [];
  var links = document.querySelectorAll('a[href]');
  for (var i = 0; i < links.length; i++) {
    var h = links[i].href || '';
    var u = '';
    var m = h.match(/[?&]q=([^&]+)/);
    if (h.indexOf('google.com/url') !== -1 && m) {
      try { u = decodeURIComponent(m[1]); } catch (e) {}
    } else if (h.indexOf('facebook.com') !== -1 && h.indexOf('google.com') === -1) {
      u = h;
    }
    if (u && u.indexOf('facebook.com') !== -1) out.push(u);
  }
  return JSON.stringify(out);
})()"""

# ================= Helpers =================

PHONE_RE = re.compile(r"(?<!\d)(?:\+?84|0)[0-9.\-\s]{9,12}(?!\d)")


def extract_phones(text):
    if not text:
        return []
    out = []
    for m in PHONE_RE.finditer(text):
        s = re.sub(r"[^0-9+]", "", m.group(0))
        if s.startswith("+84"):
            s = "0" + s[3:]
        elif s.startswith("84") and len(s) == 11:
            s = "0" + s[2:]
        if re.fullmatch(r"0\d{9}", s) and s not in out:
            out.append(s)
    return out


def clean_fb_url(u: str) -> str:
    try:
        p = urllib.parse.urlparse(u)
        if p.netloc in ("www.facebook.com", "m.facebook.com", "facebook.com") and not p.path.startswith("/search"):
            return urllib.parse.urlunparse((p.scheme, "www.facebook.com", p.path, "", "", ""))
    except Exception:
        pass
    return ""


def is_post_url(u: str) -> bool:
    return any(x in u for x in ("/posts/", "story_fbid", "/permalink.php", "/videos/", "/photo.php", "/watch", "/reel/", "video.php"))


def is_group_url(u: str) -> bool:
    return "/groups/" in u and not is_post_url(u)


def is_profile_url(u: str) -> bool:
    if is_group_url(u) or is_post_url(u):
        return False
    p = urllib.parse.urlparse(u).path
    if "/profile.php" in p:
        return True
    if p.count("/") <= 2 and len(p.strip("/")) >= 3:
        return True
    return False


def google_discover(kw: str, tabid: str) -> list:
    """Google search site:facebook.com "<kw>" ngay trong Chrome, trả list URL FB sạch."""
    q = urllib.parse.quote(f'site:facebook.com "{kw}"')
    nav(f"https://www.google.com/search?q={q}", tabid)
    wait_ready(tabid, extra=4)
    raw = js(GOOGLE_LINKS_JS, tabid).strip()
    try:
        links = json.loads(raw or "[]")
    except json.JSONDecodeError:
        links = []
    urls, seen = [], set()
    for u in links:
        cu = clean_fb_url(u)
        if cu and cu not in seen:
            seen.add(cu)
            urls.append(cu)
    return urls


def is_decoy(items) -> bool:
    """FB search trả decoy: text chứa zero-width chars hoặc link quay lại /search/."""
    for it in items:
        t = it.get("text", "")
        if any(c in t for c in "\u200b\u200c\u200d\ufeff"):
            return True
        if "/search/" in it.get("postUrl", ""):
            return True
    return False


# ================= Collectors =================

def collect_fb_search(kw: str, tabid: str, scrolls: int, max_items: int) -> list:
    q = urllib.parse.quote(kw)
    nav(f"https://www.facebook.com/search/posts/?q={q}", tabid)
    wait_ready(tabid, extra=5)
    state = js(LOGIN_CHECK_JS, tabid).strip()
    if state != "OK":
        raise RuntimeError("Facebook yêu cầu đăng nhập — đăng nhập FB trong Chrome rồi chạy lại.")
    for _ in range(scrolls):
        js("window.scrollBy(0, 1400);", tabid)
        time.sleep(1.5)
    # extract từng article có link /posts/ hoặc story_fbid (giống feed group/homepage)
    raw = js(
        r"""(function(){
  var out = [];
  var seen = {};
  var arts = document.querySelectorAll('div[role="article"]');
  for (var i = 0; i < arts.length; i++) {
    var art = arts[i];
    var t = (art.innerText || '').replace(/\s+/g, ' ').trim();
    if (t.length < 10) continue;
    var rec = {text:'', author:'', postUrl:'', time:''};
    var msg = art.querySelector('[data-ad-preview="message"]');
    rec.text = (msg ? msg.innerText : t).replace(/\s+/g, ' ').slice(0, 1500);
    var links = art.querySelectorAll('a[href]');
    for (var j = 0; j < links.length; j++) {
      var a = links[j];
      var h = a.href || '';
      var path = h.split('?')[0];
      var lt = (a.innerText || '').trim().replace(/\s+/g, ' ');
      if (!rec.postUrl && path.indexOf('/posts/') !== -1 && lt.length < 80) rec.postUrl = path;
      if (!rec.postUrl && h.indexOf('story_fbid') !== -1) rec.postUrl = h;
      if (!rec.author && lt.length > 0 && lt.length < 50 && path.indexOf('facebook.com') !== -1 &&
          path.indexOf('/search') === -1 && path.indexOf('/posts/') === -1 && path.indexOf('/hashtag/') === -1 &&
          path.indexOf('/groups/') === -1) rec.author = lt;
      if (!rec.author && lt.length > 0 && lt.length < 50 && path.indexOf('/user/') !== -1) rec.author = lt;
      if (!rec.time) {
        var al = a.getAttribute('aria-label') || '';
        if (al.length < 60 && /giờ|phút|ngày|tháng|năm|lúc|hôm qua/i.test(al)) rec.time = al;
      }
    }
    var key = rec.postUrl || (rec.author + rec.text.slice(0, 30));
    if (seen[key]) continue;
    seen[key] = 1;
    out.push(rec);
  }
  return JSON.stringify(out);
})()""",
        tabid,
    ).strip()
    try:
        items = json.loads(raw or "[]")
    except json.JSONDecodeError:
        items = []
    return items[:max_items]


def collect_page_info(url: str, tabid: str) -> dict:
    """Mở URL, lấy thông tin trang + (nếu là post) text bài. Trả dict."""
    nav(url, tabid)
    wait_ready(tabid, extra=4)
    page = json.loads(js(PAGE_JS, tabid).strip() or "{}")
    info = {
        "name": page.get("name", ""),
        "meta": page.get("meta", ""),
        "kind": page.get("kind", ""),
        "text": page.get("text", ""),
        "broken": page.get("broken", False),
        "postText": "",
        "author": "",
        "time": "",
        "login": False,
    }
    body = page.get("text", "")
    if not info["broken"] and ("Bạn phải đăng nhập" in body or "/login" in js("window.location.href", tabid)):
        info["login"] = True
    if not info["broken"] and is_post_url(url):
        try:
            post = json.loads(js(POST_PAGE_JS, tabid).strip() or "{}")
            info["postText"] = post.get("postText", "")
            info["author"] = post.get("author", "")
            info["time"] = post.get("time", "")
        except (RuntimeError, json.JSONDecodeError):
            pass
    return info


def collect_comments(tabid: str) -> list:
    for _ in range(4):
        js("window.scrollBy(0, 1200);", tabid)
        time.sleep(1)
        try:
            js(EXPAND_JS, tabid)
        except RuntimeError:
            pass
        time.sleep(1)
    raw = js(COMMENTS_JS, tabid).strip()
    try:
        return json.loads(raw or "[]")
    except json.JSONDecodeError:
        return []


# ================= Export =================

def export(rows, out_path):
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    json_path = out_path.with_suffix(".json")
    json_path.write_text(
        json.dumps(
            {"collected_at": datetime.now().isoformat(timespec="seconds"), "rows": rows},
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        import csv

        csv_path = out_path.with_suffix(".csv")
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["STT", "Từ khoá", "Loại", "Tên", "Nội dung", "SĐT", "Thời gian", "Meta", "URL"])
            for i, r in enumerate(rows, 1):
                w.writerow(
                    [i, r["keyword"], r["type"], r["name"], r["text"], " / ".join(r["phones"]), r["time"], r["meta"], r["url"]]
                )
        print(f"[!] openpyxl chưa cài — đã xuất CSV thay thế: {csv_path}")
        return csv_path

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kết quả"
    headers = ["STT", "Từ khoá", "Loại", "Tên", "Nội dung", "SĐT", "Thời gian", "Meta", "URL"]
    ws.append(headers)
    for i, r in enumerate(rows, 1):
        ws.append(
            [i, r["keyword"], r["type"], r["name"], r["text"], " / ".join(r["phones"]), r["time"], r["meta"], r["url"]]
        )
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {"A": 6, "B": 18, "C": 10, "D": 28, "E": 70, "F": 16, "G": 16, "H": 26, "I": 60}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        for c in row:
            c.number_format = "@"
    wb.save(out_path)
    return out_path


# ================= Main =================

def main():
    ap = argparse.ArgumentParser(description="Thu thập thông tin Facebook theo từ khoá (Chrome thật + AppleScript).")
    ap.add_argument("--keywords", nargs="*", help="Từ khoá cần tìm (nhiều từ cách space, nhớ quote)")
    ap.add_argument("--keywords-file", help="File chứa từ khoá, mỗi dòng một từ")
    ap.add_argument("--source", choices=["google", "fb"], default="google",
                    help="google (mặc định, bền): tìm qua Google → mở URL FB. fb: search trực tiếp trên FB (dễ bị decoy)")
    ap.add_argument("--type", choices=["all", "posts", "groups", "pages", "people"], default="all",
                    help="Lọc loại URL cần thu (mặc định all)")
    ap.add_argument("--scroll", type=int, default=8, help="Số lần scroll (fb source / trang post)")
    ap.add_argument("--limit", type=int, default=15, help="Số URL tối đa xử lý mỗi từ khoá (google source)")
    ap.add_argument("--comments", action="store_true", help="Thu thêm bình luận của các URL dạng post")
    ap.add_argument("--max-comment-posts", type=int, default=5, help="Giới hạn bài mở để lấy comment (mặc định 5)")
    ap.add_argument("--out", default=str(DEFAULT_OUT), help=f"File xuất (mặc định {DEFAULT_OUT})")
    ap.add_argument("--no-raise", action="store_true",
                    help="Không kéo Chrome lên trước (có thể khiến FB không render lazy content)")
    ap.add_argument("--check", action="store_true", help="Chỉ kiểm tra môi trường: Chrome + quyền JS + đăng nhập FB")
    args = ap.parse_args()

    if args.check:
        ensure_chrome()
        tabid = ensure_tab()
        if not args.no_raise:
            raise_chrome()
        nav("https://www.facebook.com/", tabid)
        wait_ready(tabid, extra=3)
        state = js(LOGIN_CHECK_JS, tabid).strip()
        if state == "OK":
            print("OK — Chrome + Facebook OK, sẵn sàng thu thập.")
        else:
            print("Chrome OK nhưng Facebook báo cần đăng nhập. Đăng nhập FB trong Chrome rồi chạy lại.")
        return

    keywords = list(args.keywords or [])
    if args.keywords_file:
        keywords += [l.strip() for l in Path(args.keywords_file).read_text(encoding="utf-8").splitlines() if l.strip()]
    if not keywords:
        ap.error('Cần ít nhất một từ khoá: --keywords "từ khoá" hoặc --keywords-file file.txt')

    ensure_chrome()
    tabid = ensure_tab()
    prev_app = ""
    if not args.no_raise:
        prev_app = frontmost_app()
        raise_chrome()
    print(f"[*] Tab #{tabid}, {len(keywords)} từ khoá, source={args.source}, type={args.type}...")
    rows = []
    seen_urls = set()

    try:
        for kw in keywords:
            try:
                if args.source == "fb":
                    items = collect_fb_search(kw, tabid, args.scroll, args.limit)
                    if is_decoy(items):
                        print(f"[!] '{kw}': FB trả decoy (chống automation). Dùng --source google sẽ bền hơn.")
                    for it in items:
                        url = it.get("postUrl", "")
                        key = "post|" + url
                        if key in seen_urls:
                            continue
                        seen_urls.add(key)
                        rows.append({
                            "keyword": kw, "type": "post", "name": it.get("author", ""),
                            "text": it.get("text", ""), "phones": extract_phones(it.get("text", "")),
                            "time": it.get("time", ""), "meta": "", "url": url,
                        })
                    print(f"[+] '{kw}': {len(items)} bài (fb source)")
                else:
                    urls = google_discover(kw, tabid)
                    filtered = [u for u in urls if args.type == "all" or
                                (args.type == "groups" and is_group_url(u)) or
                                (args.type == "posts" and is_post_url(u)) or
                                (args.type == "pages" and is_profile_url(u) and "/groups/" not in u and not is_post_url(u)) or
                                (args.type == "people" and is_profile_url(u))]
                    if not filtered and urls:
                        filtered = urls[: args.limit]  # fallback: lấy hết nếu filter quá khắt
                    filtered = filtered[: args.limit]
                    print(f"[+] '{kw}': {len(urls)} URL FB tìm thấy, xử lý {len(filtered)}")
                    for u in filtered:
                        key = "url|" + u
                        if key in seen_urls:
                            continue
                        seen_urls.add(key)
                        try:
                            info = collect_page_info(u, tabid)
                        except RuntimeError as e:
                            print(f"    [!] {u[:70]}: {e}")
                            continue
                        if info["broken"]:
                            rows.append({
                                "keyword": kw, "type": "broken", "name": u.split("/")[-2] or u,
                                "text": "Trang này không hiển thị (link hỏng/đã gỡ)", "phones": [],
                                "time": "", "meta": "", "url": u,
                            })
                            print(f"    - {u[:70]} (broken)")
                            continue
                        if info["login"]:
                            rows.append({
                                "keyword": kw, "type": "login", "name": info["name"],
                                "text": "Cần đăng nhập mới xem được", "phones": [],
                                "time": "", "meta": info["meta"], "url": u,
                            })
                            continue
                        typ = "post" if (is_post_url(u) and info["postText"]) else ("group" if is_group_url(u) else "page")
                        name = info["author"] or info["name"]
                        text = info["postText"] or info["text"]
                        rows.append({
                            "keyword": kw, "type": typ, "name": name, "text": text[:1500],
                            "phones": extract_phones(text), "time": info["time"], "meta": info["meta"], "url": u,
                        })
                        print(f"    - {typ:6s} | {name[:35]:35s} | {u[:55]}")
                        # comments
                        if args.comments and is_post_url(u) and info["postText"] and info["author"]:
                            try:
                                cs = collect_comments(tabid)
                            except RuntimeError as e:
                                print(f"      [!] comments {u[:60]}: {e}")
                                continue
                            for c in cs:
                                rows.append({
                                    "keyword": kw, "type": "comment", "name": c.get("author", ""),
                                    "text": c.get("text", ""), "phones": extract_phones(c.get("text", "")),
                                    "time": "", "meta": info["author"], "url": u,
                                })
                            if cs:
                                print(f"      ↳ {len(cs)} comment")
                            time.sleep(1.5)
            except RuntimeError as e:
                print(f"[!] Lỗi từ khoá '{kw}': {e}")
            time.sleep(1)
    finally:
        if not args.no_raise and prev_app and prev_app.lower() != "google chrome":
            try:
                osa(f'tell application "{prev_app}" to activate', timeout=30)
            except RuntimeError:
                pass

    if not rows:
        print("[!] Không thu được dữ liệu. Kiểm tra: đăng nhập FB, từ khoá, hoặc chạy --check.")
        return
    out = export(rows, args.out)
    print(f"[✓] Xong: {len(rows)} dòng → {out}")
    print(f"    JSON: {Path(out).with_suffix('.json')}")
    for r in rows[:3]:
        print(f"    - {r['type']} | {r['name'][:40]} | {r['url'][:70]}")


if __name__ == "__main__":
    main()
