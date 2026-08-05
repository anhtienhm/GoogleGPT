#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FB Group Watch — theo dõi nhóm Facebook hàng ngày.

Điều khiển Chrome thật (đã đăng nhập FB) qua AppleScript + execute javascript.
Spec (theo yêu cầu):
  - Mỗi nhóm: quét feed → lấy tối đa 20 bài có nhiều bình luận
  - Mở từng bài → đếm bình luận → thu comment của bài đạt ngưỡng
  - Lọc comment: trong vòng 24 giờ + liên quan SEO
  - Tổng hợp link bài có comment đạt lọc → links.txt → push lên repo (git)
  - Xuất báo cáo Excel + JSON

Usage:
  python3 fb-group-watch.py --groups "<url1>" "<url2>"
  python3 fb-group-watch.py --groups "<url1>" --max-posts 20 --hours 24
"""
import argparse
import json
import re
import subprocess
import tempfile
import time
from datetime import datetime, timedelta
from pathlib import Path

OUT_DIR = Path.home() / "Facebook"
DEFAULT_OUT = OUT_DIR / f"fb_group_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
DEFAULT_LINKS = Path.home() / "GoogleGPT" / "links.txt"

# Từ khoá SEO (comment phải chứa ít nhất 1)
SEO_RE = re.compile(
    r"(?i)(\bseo\b|backlink|textlink|text link|link báo|báo tài chính|báo lớn|google|từ khóa|từ khoá|"
    r"keyword|ranking|thứ hạng|lên top|webmaster|gmb|map listing|domain|crawl|traffic|anchor|guest post|"
    r"pr báo|dofollow|nofollow|web 2\.0|pbn|wordpress|website|lên google|index nhanh|tăng ranking)"
)

# ================= AppleScript helpers =================

def osa(script: str, timeout: int = 90) -> str:
    tmp = Path(tempfile.mktemp(suffix=".applescript"))
    tmp.write_text(script, encoding="utf-8")
    try:
        try:
            r = subprocess.run(["osascript", str(tmp)], capture_output=True, text=True, timeout=timeout)
        except subprocess.TimeoutExpired:
            raise RuntimeError(f"osascript timeout ({timeout}s) — trang quá tải hoặc JS bị treo")
    finally:
        try:
            tmp.unlink()
        except OSError:
            pass
    if r.returncode != 0:
        err = (r.stderr or "").strip()
        if "-1743" in err:
            raise RuntimeError(
                "macOS chặn quyền điều khiển Chrome. Vào System Settings → Privacy & Security → "
                "Automation, bật quyền điều khiển Google Chrome cho app đang chạy (Hermes/Terminal)."
            )
        if "not allowed" in err.lower() or "not authorized" in err.lower():
            raise RuntimeError(
                'Chrome chưa bật "Allow JavaScript from Apple Events": '
                "menu Chrome → View → Developer → Allow JavaScript from Apple Events.\n" + err
            )
        raise RuntimeError(f"osascript thất bại: {err}")
    return r.stdout


def frontmost_app() -> str:
    try:
        return osa(
            'tell application "System Events"\n'
            "    set f to name of first application process whose frontmost is true\n"
            "    return f\n"
            "end tell",
            timeout=30,
        ).strip()
    except RuntimeError:
        return ""


def ensure_chrome() -> None:
    try:
        osa('tell application "Google Chrome" to count windows', timeout=30)
    except RuntimeError:
        subprocess.run(["open", "-a", "Google Chrome"], timeout=30)
        time.sleep(5)


def raise_chrome() -> None:
    osa(
        'tell application "Google Chrome"\n'
        "    if (count of windows) > 0 then\n"
        "        set index of front window to 1\n"
        "    end if\n"
        "end tell\n"
        'tell application "Google Chrome" to activate',
        timeout=30,
    )


def ensure_tab() -> str:
    out = osa(
        'tell application "Google Chrome"\n'
        "    tell front window\n"
        '        make new tab with properties {URL:"about:blank"}\n'
        "        return (id of last tab)\n"
        "    end tell\n"
        "end tell",
        timeout=30,
    )
    return out.strip()


def activate_tab(tabid: str) -> None:
    osa(
        f'tell application "Google Chrome"\n'
        "    tell front window\n"
        "        set i to 0\n"
        "        repeat with t in tabs\n"
        "            set i to i + 1\n"
        f'            if ((id of t) as text) = "{tabid}" then\n'
        "                set active tab index to i\n"
        "                exit repeat\n"
        "            end if\n"
        "        end repeat\n"
        "    end tell\n"
        "end tell",
        timeout=30,
    )


def nav(url: str, tabid: str) -> None:
    osa(f'tell application "Google Chrome" to set URL of tab id {tabid} of front window to "{url}"', timeout=60)


def js(code: str, tabid: str, retries: int = 1) -> str:
    activate_tab(tabid)
    esc = code.replace("\\", "\\\\").replace('"', '\\"')
    script = (
        'tell application "Google Chrome"\n'
        "    tell active tab of front window\n"
        f'        set r to execute javascript "{esc}"\n'
        "    end tell\n"
        "end tell\n"
        "return r"
    )
    for attempt in range(retries + 1):
        try:
            return osa(script, timeout=35)
        except RuntimeError:
            if attempt >= retries:
                raise
            time.sleep(2)


def wait_ready(tabid: str, extra: float = 3.0, tries: int = 15) -> None:
    for _ in range(tries):
        try:
            if js("document.readyState", tabid).strip() == "complete":
                break
        except Exception:
            pass
        time.sleep(1)
    time.sleep(extra)


# ================= JS snippets =================

LOGIN_CHECK_JS = r"""(function(){
  if (document.querySelector('input[name="email"], input[name="pass"], #loginbutton')) {
    return 'LOGIN_REQUIRED';
  }
  return 'OK';
})()"""

GROUP_INFO_JS = r"""(function(){
  var out = {name:'', members:''};
  var body = (document.body ? document.body.innerText : '') || '';
  body = body.replace(/\s+/g, ' ');
  var t = (document.title || '').replace(/^\(\d+\)\s*/, '').replace(/\s*\|\s*Facebook\s*$/, '').trim();
  out.name = t;
  var m = body.match(/([\d.,]+[KkMm]?\s*thành viên)/);
  if (m) out.members = m[1];
  return JSON.stringify(out);
})()"""

FEED_POSTS_JS = r"""(function(){
  var out = [];
  var seen = {};
  var arts = document.querySelectorAll('div[role="article"]');
  for (var i = 0; i < arts.length; i++) {
    var art = arts[i];
    var t = (art.innerText || '').replace(/\s+/g, ' ').trim();
    if (t.length < 10) continue;
    var rec = {author:'', text:'', postUrl:'', time:'', reactions:''};
    var msg = art.querySelector('[data-ad-preview="message"]');
    rec.text = (msg ? msg.innerText : t).replace(/\s+/g, ' ').slice(0, 1500);
    var links = art.querySelectorAll('a[href]');
    for (var j = 0; j < links.length; j++) {
      var a = links[j];
      var h = a.href || '';
      var path = h.split('?')[0];
      var lt = (a.innerText || '').trim().replace(/\s+/g, ' ');
      if (!rec.postUrl && path.indexOf('/posts/') !== -1 && lt.length < 80) rec.postUrl = path;
      if (!rec.postUrl && h.indexOf('story_fbid') !== -1) rec.postUrl = h;
      if (!rec.author && lt.length > 0 && lt.length < 50 && path.indexOf('facebook.com') !== -1 &&
          path.indexOf('/search') === -1 && path.indexOf('/posts/') === -1 && path.indexOf('/hashtag/') === -1 &&
          path.indexOf('/groups/') === -1) rec.author = lt;
      if (!rec.author && lt.length > 0 && lt.length < 50 && path.indexOf('/user/') !== -1) rec.author = lt;
      if (!rec.time) {
        var al = a.getAttribute('aria-label') || '';
        if (al.length < 80 && /giờ|phút|ngày|tháng|năm|lúc|hôm qua|hôm nay|vừa xong/i.test(al)) rec.time = al;
      }
    }
    var labelled = art.querySelectorAll('[aria-label]');
    for (var j = 0; j < labelled.length; j++) {
      var l = labelled[j].getAttribute('aria-label') || '';
      var m = l.match(/^([\d.,]+[KkMm]?)\s*(?:người|thành viên)?\s*(?:đã bày tỏ cảm xúc|thích)/);
      if (m && !rec.reactions) rec.reactions = m[1];
    }
    if (!rec.reactions) {
      var btns = art.querySelectorAll('div[role="button"]');
      for (var j = 0; j < btns.length; j++) {
        var bt = (btns[j].innerText || '').trim();
        if (/^[\d.,]+[KkMm]?$/.test(bt) && !rec.reactions) rec.reactions = bt;
      }
    }
    var key = rec.postUrl || (rec.author + rec.text.slice(0, 30));
    if (seen[key]) continue;
    seen[key] = 1;
    out.push(rec);
  }
  return JSON.stringify(out);
})()"""

POST_DETAIL_JS = r"""(function(){
  var out = {comments: 0, title: '', reactions: '', broken: false};
  var body = (document.body ? document.body.innerText : '') || '';
  body = body.replace(/\s+/g, ' ');
  if (body.indexOf('Trang này không hiển thị') !== -1) out.broken = true;
  out.comments = document.querySelectorAll('[aria-label^="Bình luận dưới tên"], [aria-label^="Comment by"]').length;
  var t = (document.title || '').replace(/^\(\d+\)\s*/, '').replace(/\s*\|\s*Facebook\s*$/, '').trim();
  var parts = t.split(' | ');
  out.title = parts.length > 1 ? parts.slice(1).join(' | ') : t;
  var m = body.match(/([\d.,]+[KkMm]?\s*người)\s*đã bày tỏ cảm xúc/);
  if (m) out.reactions = m[1];
  return JSON.stringify(out);
})()"""

EXPAND_JS = r"""(function(){
  var n = 0;
  var btns = document.querySelectorAll('div[role="button"], span[role="button"]');
  for (var i = 0; i < btns.length; i++) {
    var t = (btns[i].innerText || '').trim();
    if (t.length < 50 && /xem thêm|see more|hiển thị thêm|bình luận trước|phản hồi trước/i.test(t)) {
      try { btns[i].click(); n++; } catch (e) {}
    }
  }
  return n;
})()"""

# Comment: bỏ article 0 (bài gốc), kèm time label "Bình luận dưới tên X vào N giờ trước"
COMMENTS_JS = r"""(function(){
  var out = [];
  var seen = {};
  var arts = document.querySelectorAll('div[role="article"]');
  for (var i = 1; i < arts.length; i++) {
    var art = arts[i];
    var t = (art.innerText || '').replace(/\s+/g, ' ').trim();
    if (t.length < 5) continue;
    var author = '';
    var links = art.querySelectorAll('a[href]');
    for (var j = 0; j < links.length; j++) {
      var h = (links[j].href || '').split('?')[0];
      var lt = (links[j].innerText || '').trim().replace(/\s+/g, ' ');
      if (!author && lt.length > 0 && lt.length < 50 && h.indexOf('facebook.com') !== -1 &&
          h.indexOf('/search') === -1 && h.indexOf('/groups/') === -1 && h.indexOf('/posts/') === -1 &&
          h.indexOf('/hashtag/') === -1 && h.indexOf('/videos/') === -1) author = lt;
      if (!author && lt.length > 0 && lt.length < 50 && h.indexOf('/user/') !== -1) author = lt;
    }
    if (!author) continue;
    var key = author + '|' + t.slice(0, 40);
    if (seen[key]) continue;
    seen[key] = 1;
    var tl = '';
    var tls = art.querySelectorAll('[aria-label^="Bình luận dưới tên"], [aria-label^="Comment by"]');
    if (tls.length) tl = tls[0].getAttribute('aria-label') || '';
    out.push({author: author, text: t.slice(0, 1000), time: tl});
  }
  return JSON.stringify(out);
})()"""

# ================= Helpers =================

PHONE_RE = re.compile(r"(?<!\d)(?:\+?84|0)[0-9.\-\s]{9,12}(?!\d)")


def extract_phones(text):
    if not text:
        return []
    out = []
    for m in PHONE_RE.finditer(text):
        s = re.sub(r"[^0-9+]", "", m.group(0))
        if s.startswith("+84"):
            s = "0" + s[3:]
        elif s.startswith("84") and len(s) == 11:
            s = "0" + s[2:]
        if re.fullmatch(r"0\d{9}", s) and s not in out:
            out.append(s)
    return out


def parse_fb_time(s: str, now: datetime) -> datetime | None:
    if not s:
        return None
    s = s.strip()
    m = re.search(r"(\d{1,2})\s+Tháng\s+(\d{1,2}),\s+(\d{4})\s+lúc\s+(\d{1,2}):(\d{2})", s)
    if m:
        try:
            return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)), int(m.group(4)), int(m.group(5)))
        except ValueError:
            return None
    m = re.search(r"Hôm nay\s+lúc\s+(\d{1,2}):(\d{2})", s)
    if m:
        return now.replace(hour=int(m.group(1)), minute=int(m.group(2)), second=0, microsecond=0)
    m = re.search(r"Hôm qua\s+lúc\s+(\d{1,2}):(\d{2})", s)
    if m:
        return (now - timedelta(days=1)).replace(hour=int(m.group(1)), minute=int(m.group(2)), second=0, microsecond=0)
    m = re.search(r"(\d+)\s+giờ", s)
    if m:
        return now - timedelta(hours=int(m.group(1)))
    m = re.search(r"(\d+)\s+phút", s)
    if m:
        return now - timedelta(minutes=int(m.group(1)))
    m = re.search(r"(\d+)\s+ngày", s)
    if m:
        return now - timedelta(days=int(m.group(1)))
    if "vừa xong" in s:
        return now - timedelta(minutes=1)
    return None


def parse_comment_time(raw: str, now: datetime) -> datetime | None:
    """'Bình luận dưới tên X vào 2 giờ trước' / 'vừa xong' / 'lúc HH:MM'."""
    if not raw:
        return None
    m = re.search(r"vào\s+(?:lúc\s+)?(\d+)\s*(phút|giờ|ngày|tuần)\s*trước", raw)
    if m:
        n = int(m.group(1))
        unit = m.group(2)
        if unit == "phút":
            return now - timedelta(minutes=n)
        if unit == "giờ":
            return now - timedelta(hours=n)
        if unit == "ngày":
            return now - timedelta(days=n)
        if unit == "tuần":
            return now - timedelta(weeks=n)
    if "vừa xong" in raw:
        return now - timedelta(minutes=1)
    m = re.search(r"lúc\s+(\d{1,2}):(\d{2})", raw)
    if m:
        return now.replace(hour=int(m.group(1)), minute=int(m.group(2)), second=0, microsecond=0)
    return None


# ================= Collectors =================

def collect_group_feed(group_url: str, tabid: str, feed_scrolls: int) -> tuple:
    nav(group_url, tabid)
    wait_ready(tabid, extra=5)
    state = js(LOGIN_CHECK_JS, tabid).strip()
    if state != "OK":
        raise RuntimeError("Facebook yêu cầu đăng nhập — đăng nhập FB trong Chrome rồi chạy lại.")
    try:
        gi = json.loads(js(GROUP_INFO_JS, tabid).strip() or "{}")
    except json.JSONDecodeError:
        gi = {}
    time.sleep(2)
    js("window.scrollBy(0, 800);", tabid)
    time.sleep(2)
    all_posts = []
    for i in range(feed_scrolls):
        js("window.scrollBy(0, 1600);", tabid)
        time.sleep(1.8)
        try:
            js(EXPAND_JS, tabid)
        except RuntimeError:
            pass
        time.sleep(0.8)
        # Feed bị virtualize — bài chỉ tồn tại trong DOM khi ở viewport.
        # Phải extract DỌC ĐƯỜNG scroll, không extract 1 lần cuối.
        if i % 3 == 2 or i == feed_scrolls - 1:
            try:
                raw = js(FEED_POSTS_JS, tabid).strip()
                all_posts.extend(json.loads(raw or "[]"))
            except (RuntimeError, json.JSONDecodeError):
                pass
    posts = []
    seen = set()
    for p in all_posts:
        key = p.get("postUrl") or (p.get("author", "") + p.get("text", "")[:30])
        if key in seen:
            continue
        seen.add(key)
        posts.append(p)
    return gi.get("name", ""), gi.get("members", ""), posts


def get_post_detail(post_url: str, tabid: str) -> dict:
    nav(post_url, tabid)
    wait_ready(tabid, extra=3)
    raw = js(POST_DETAIL_JS, tabid).strip()
    try:
        return json.loads(raw or "{}")
    except json.JSONDecodeError:
        return {}


def collect_comments(tabid: str) -> list:
    for _ in range(4):
        js("window.scrollBy(0, 1200);", tabid)
        time.sleep(1)
        try:
            js(EXPAND_JS, tabid)
        except RuntimeError:
            pass
        time.sleep(1)
    raw = js(COMMENTS_JS, tabid).strip()
    try:
        return json.loads(raw or "[]")
    except json.JSONDecodeError:
        return []


# ================= Export =================

def export(rows, out_path):
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    json_path = out_path.with_suffix(".json")
    json_path.write_text(
        json.dumps(
            {
                "collected_at": datetime.now().isoformat(timespec="seconds"),
                "rows": [{k: v for k, v in r.items() if k != "_dt"} for r in rows],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        import csv

        csv_path = out_path.with_suffix(".csv")
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["STT", "Nhóm", "Ngày thu", "Loại", "Tác giả", "Nội dung", "SĐT", "Thời gian", "Bình luận", "Reactions", "URL"])
            for i, r in enumerate(rows, 1):
                w.writerow(
                    [i, r["group"], r["collected"], r["type"], r["author"], r["text"], " / ".join(r["phones"]),
                     r["time"], r["comments"], r["reactions"], r["url"]]
                )
        print(f"[!] openpyxl chưa cài — đã xuất CSV thay thế: {csv_path}")
        return csv_path

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kết quả"
    headers = ["STT", "Nhóm", "Ngày thu", "Loại", "Tác giả", "Nội dung", "SĐT", "Thời gian", "Bình luận", "Reactions", "URL"]
    ws.append(headers)
    for i, r in enumerate(rows, 1):
        ws.append(
            [i, r["group"], r["collected"], r["type"], r["author"], r["text"], " / ".join(r["phones"]),
             r["time"], r["comments"], r["reactions"], r["url"]]
        )
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {"A": 6, "B": 30, "C": 18, "D": 9, "E": 26, "F": 70, "G": 16, "H": 24, "I": 10, "J": 12, "K": 55}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
        for c in row:
            c.number_format = "@"
    wb.save(out_path)
    return out_path


# ================= Main =================

def main():
    ap = argparse.ArgumentParser(description="Theo dõi nhóm Facebook: 20 bài nhiều bình luận, comment 24h liên quan SEO.")
    ap.add_argument("--groups", nargs="+", required=True, help="URL các group (bắt buộc)")
    ap.add_argument("--max-posts", type=int, default=10, help="Số bài mỗi nhóm xử lý (mặc định 10)")
    ap.add_argument("--min-comments", type=int, default=2, help="Bài phải có ≥ N bình luận mới thu comment (mặc định 2)")
    ap.add_argument("--hours", type=int, default=24, help="Lọc comment trong N giờ gần đây (mặc định 24)")
    ap.add_argument("--feed-scrolls", type=int, default=18, help="Số lần scroll feed group (mặc định 18)")
    ap.add_argument("--out", default=str(DEFAULT_OUT), help=f"File xuất (mặc định {DEFAULT_OUT})")
    ap.add_argument(
        "--links-out",
        default=str(DEFAULT_LINKS),
        help=f"File links.txt tổng hợp (mặc định {DEFAULT_LINKS})",
    )
    ap.add_argument("--no-raise", action="store_true", help="Không kéo Chrome lên trước")
    args = ap.parse_args()

    ensure_chrome()
    tabid = ensure_tab()
    prev_app = ""
    if not args.no_raise:
        prev_app = frontmost_app()
        raise_chrome()
    now = datetime.now()
    print(f"[*] Tab #{tabid}, {len(args.groups)} group, max {args.max_posts} bài/nhóm, comment {args.hours}h + SEO...")

    rows = []
    try:
        for gurl in args.groups:
            gname = gurl.rstrip("/").split("/")[-1]
            try:
                name, members, posts = collect_group_feed(gurl, tabid, args.feed_scrolls)
            except RuntimeError as e:
                print(f"[!] Group {gurl}: {e}")
                continue
            print(f"[+] {name or gname} ({members}): {len(posts)} bài trong feed, xử lý tối đa {args.max_posts}")
            for p in posts[: args.max_posts]:
                url = p.get("postUrl", "")
                if not url:
                    continue
                try:
                    det = get_post_detail(url, tabid)
                except RuntimeError as e:
                    print(f"    [!] {url[:70]}: {e}")
                    continue
                if det.get("broken"):
                    print(f"    - (broken) {url[:70]}")
                    continue
                comments_n = det.get("comments", 0)
                p["comments"] = comments_n
                p["title"] = det.get("title", "") or p.get("text", "")
                p["reactions"] = det.get("reactions", "") or p.get("reactions", "")
                p["_dt"] = parse_fb_time(p.get("time", ""), now)
                qualifying = []
                if comments_n >= args.min_comments:
                    try:
                        cs = collect_comments(tabid)
                    except RuntimeError:
                        cs = []
                    for c in cs:
                        cdt = parse_comment_time(c.get("time", ""), now)
                        fresh = (not cdt) or (now - cdt) <= timedelta(hours=args.hours)
                        seo = bool(SEO_RE.search(c.get("text", "")))
                        if fresh and seo:
                            qualifying.append(c)
                p["qualifying"] = qualifying
                mark = f" ✓ {len(qualifying)} cmt đạt" if qualifying else ""
                print(f"    - {p.get('author','')[:20]:20s} | {comments_n:3d} cmt | {p.get('time','')[:32]}{mark}")
                time.sleep(0.8)
            # rows: post (context) + comment đạt lọc
            for p in posts[: args.max_posts]:
                if not p.get("postUrl"):
                    continue
                rows.append({
                    "group": name or gname,
                    "collected": now.strftime("%Y-%m-%d %H:%M"),
                    "type": "post",
                    "author": p.get("author", ""),
                    "text": p.get("title", "")[:1500],
                    "phones": extract_phones(p.get("title", "")),
                    "time": p.get("time", ""),
                    "comments": p.get("comments", 0),
                    "reactions": p.get("reactions", ""),
                    "url": p.get("postUrl", ""),
                    "_dt": p.get("_dt"),
                })
                for c in p.get("qualifying", []):
                    rows.append({
                        "group": name or gname,
                        "collected": now.strftime("%Y-%m-%d %H:%M"),
                        "type": "comment",
                        "author": c.get("author", ""),
                        "text": c.get("text", ""),
                        "phones": extract_phones(c.get("text", "")),
                        "time": c.get("time", ""),
                        "comments": 0,
                        "reactions": "",
                        "url": p.get("postUrl", ""),
                        "_dt": p.get("_dt"),
                    })
    finally:
        if not args.no_raise and prev_app and prev_app.lower() != "google chrome":
            try:
                osa(f'tell application "{prev_app}" to activate', timeout=30)
            except RuntimeError:
                pass

    if not rows:
        print("[!] Không thu được dữ liệu. Kiểm tra đăng nhập FB hoặc chạy --no-raise nếu cần.")
        return

    out = export(rows, args.out)
    print(f"\n[✓] Báo cáo: {len(rows)} dòng → {out}")
    print(f"    JSON: {Path(out).with_suffix('.json')}")

    # links.txt: bài có ≥1 comment đạt lọc, sort theo tổng bình luận giảm dần
    post_rows = [r for r in rows if r["type"] == "post" and r["comments"] > 0]
    has_qual = [r for r in rows if r["type"] == "comment"]
    qual_urls = {r["url"] for r in has_qual}
    links = [r["url"] for r in post_rows if r["url"] in qual_urls]
    links = list(dict.fromkeys(links))  # giữ thứ tự, dedupe
    links_path = Path(args.links_out)
    if links:
        links_path.parent.mkdir(parents=True, exist_ok=True)
        links_path.write_text("\n".join(links) + "\n", encoding="utf-8")
        print(f"[✓] links.txt: {len(links)} link (bài có comment 24h + SEO) → {links_path}")
    else:
        print("[!] Không có bài nào có comment đạt lọc → links.txt KHÔNG ghi (tránh ghi đè file cũ).")

    # top 10 summary
    top = sorted(post_rows, key=lambda r: (r["comments"], r.get("_dt") or datetime.min), reverse=True)[:10]
    print("\n=== TOP 10 NHIỀU BÌNH LUẬN ===")
    for r in top:
        print(f"  {r['comments']:3d} cmt | {r['group'][:26]:26s} | {r['author'][:20]:20s} | {r['url'][:60]}")


if __name__ == "__main__":
    main()
