# -*- coding: utf-8 -*-
"""
exporter.py — Chuẩn hoá dữ liệu lead & ghi Excel.

Tách khỏi app.py để scraper không phải biết gì về định dạng file đầu ra.
Không dùng pandas: dữ liệu là list[dict] nhỏ, openpyxl thuần nhanh hơn và
không sinh NaN.

Public API:
    COLUMNS               # thứ tự cột chuẩn
    norm_phone(v)         # +84xxx / 84xxx / 0xxx  -> 0xxx
    norm_name_key(v)      # khoá so khớp tên, xử lý đúng chữ 'đ'
    apply_phone_info(row, info)   # gán kết quả FBnumber vào 1 row
    merge_rows(rows)      # dedupe theo uid > url > tên, gộp comment
    save_excel(rows, path)
"""

from __future__ import annotations

import os
import re
import unicodedata
from typing import Any, Dict, Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------- #
# CONSTANTS
# --------------------------------------------------------------------------- #

COLUMNS = [
    "Ngày tìm", "Tên KH", "SĐT", "Nhà mạng", "SĐT 2", "Nhà mạng 2",
    "Số lần tìm", "Email", "Emails", "Giới tính", "Ngày sinh",
    "Location", "Comment", "Link bài viết", "Facebook", "Zalo",
]

COL_WIDTHS = [19, 26, 14, 12, 14, 12, 10, 24, 24, 9, 12, 26, 50, 50, 44, 30]
PHONE_COLS = ("SĐT", "SĐT 2")          # phải ghi dạng text, giữ số 0 đầu
WRAP_COLS = ("Comment", "Location")

# Placeholder coi như "chưa có tên thật"
PLACEHOLDER_NAMES = {"", "n/a", "người dùng facebook", "nguoi dung facebook"}
NULLS = {"", "n/a", "na", "none", "null", "nan", "-"}

_RE_UID_NAME = re.compile(r"^\s*UID\s*[:\-]?\s*(\d{5,})\s*$", re.I)
_RE_NON_ALNUM = re.compile(r"[^a-z0-9]+")

COMMENT_JOIN = " | "


# --------------------------------------------------------------------------- #
# NORMALIZERS (pure functions)
# --------------------------------------------------------------------------- #

def _s(value: Any) -> str:
    """Về str đã strip; coi None/nan/N/A là rỗng."""
    s = "" if value is None else str(value).strip()
    return "" if s.lower() in NULLS else s


def norm_phone(value: Any) -> str:
    """+84776791717 / 84776791717 / 0776791717 -> 0776791717"""
    d = re.sub(r"\D", "", _s(value))
    if d.startswith("84") and len(d) >= 11:
        d = "0" + d[2:]
    elif d and not d.startswith("0") and len(d) == 9:
        d = "0" + d
    return d if 10 <= len(d) <= 11 else ""


def norm_name_key(value: Any) -> str:
    """Khoá so khớp tên: bỏ dấu, lowercase, bỏ ký tự đặc biệt.

    KHÁC với NFKD + ASCII-ignore: cách đó xoá sạch chữ 'đ' (Đạt -> at).
    Ở đây map đ->d trước khi bỏ dấu.
    """
    s = _s(value).replace("Đ", "D").replace("đ", "d")
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return _RE_NON_ALNUM.sub("", s.lower())


def is_placeholder_name(value: Any) -> bool:
    """True nếu 'Tên KH' chưa phải tên nick thật (rỗng / UID 100... / mặc định)."""
    s = _s(value)
    return s.lower() in PLACEHOLDER_NAMES or bool(_RE_UID_NAME.match(s))


def uid_from_url(url: Any) -> str:
    """Lấy UID từ profile URL (số hoặc pfbid). Giữ đồng bộ với extract_uid_from_url ở app.py."""
    u = _s(url)
    for rx in (r"pfbid[A-Za-z0-9]{10,}", r"/people/[^/]+/(\d+|pfbid[A-Za-z0-9]+)",
               r"profile\.php\?id=(\d+)", r"/user/(\d+)", r"facebook\.com/(\d{10,})(?:[/?#]|$)"):
        m = re.search(rx, u)
        if m:
            return m.group(1) if m.lastindex else m.group(0)
    return ""


# --------------------------------------------------------------------------- #
# FBNUMBER -> ROW
# --------------------------------------------------------------------------- #

def apply_phone_info(row: Dict[str, Any], info: Dict[str, Any], uid: str = "",
                     real_name: str = "") -> bool:
    """Gán kết quả tra cứu FBnumber vào row. Trả về True nếu có ít nhất 1 SĐT.

    KHÔNG ghi đè SĐT nếu row đã có (từ comment text).
    Gộp 2 khối code trùng nhau ở app.py (priority-2 match và positional fallback).
    SĐT chính và SĐT 2 nằm ở 2 cột riêng — KHÔNG join vào cùng 1 ô.
    """
    phones = [norm_phone(info.get("number")), norm_phone(info.get("number2"))]
    phones = [p for p in dict.fromkeys(phones) if p]      # dedupe, giữ thứ tự
    if not phones:
        return False

    # Chỉ ghi SĐT nếu chưa có (tránh ghi đè SĐT từ comment text)
    if not _s(row.get("SĐT")):
        row["SĐT"] = phones[0]
    if not _s(row.get("Nhà mạng")):
        row["Nhà mạng"] = _s(info.get("numberProvider"))
    if len(phones) > 1:
        if not _s(row.get("SĐT 2")):
            row["SĐT 2"] = phones[1]
        if not _s(row.get("Nhà mạng 2")):
            row["Nhà mạng 2"] = _s(info.get("number2Provider"))

    for src, col in (("location", "Location"), ("gender", "Giới tính"),
                     ("birthday", "Ngày sinh"), ("email", "Email")):
        val = _s(info.get(src))
        if val and not _s(row.get(col)):
            row[col] = val

    emails = info.get("emails") or []
    if isinstance(emails, (list, tuple)) and len(emails) > 1:
        row["Emails"] = ", ".join(_s(e) for e in emails[1:] if _s(e))

    # Backfill tên nick thật + link FB từ uid_map của GraphQL interceptor
    if uid:
        row["_uid"] = uid
        if not _s(row.get("Facebook")):
            row["Facebook"] = f"https://www.facebook.com/{uid}"
    if real_name and is_placeholder_name(row.get("Tên KH")):
        row["Tên KH"] = _s(real_name)

    return True


# --------------------------------------------------------------------------- #
# MERGE / DEDUPE
# --------------------------------------------------------------------------- #

def merge_phone_lookup(
    all_data: List[Dict[str, Any]],
    uid_map: Dict[str, str],
    uid_to_phone: Dict[str, Dict[str, Any]],
    allow_positional: bool = True,
    verbose: bool = True,
) -> Dict[str, int]:
    """Gán kết quả FBnumber vào leads. Hàm thuần -> test được không cần Selenium.

    Thứ tự ưu tiên:
      1. UID lấy từ href       (chắc chắn)
      2. Tên nick chuẩn hoá    (gần chắc chắn)
      3. Vị trí, chỉ khi số lượng khớp 1-1 (fallback, có cảnh báo)

    Mutate all_data tại chỗ. Trả về dict thống kê để caller in log.
    """
    stats = {"uid": 0, "name": 0, "positional": 0, "pending": 0}
    if not uid_to_phone:
        return stats

    # Index tên -> uid: dựng 1 lần, O(n+m) thay vì O(n*m)
    name_to_uid: Dict[str, str] = {}
    for uid, name in uid_map.items():
        if uid in uid_to_phone:
            name_to_uid.setdefault(norm_name_key(name), uid)

    for row in all_data:
        row_uid = _s(row.get("_uid"))
        matched = row_uid if row_uid in uid_to_phone else None
        by_uid = matched is not None
        if not matched:
            matched = name_to_uid.get(norm_name_key(row.get("Tên KH")))
        if matched and apply_phone_info(row, uid_to_phone[matched], uid=matched,
                                        real_name=uid_map.get(matched, "")):
            stats["uid" if by_uid else "name"] += 1

    if verbose:
        print(f"-> Match SĐT: {stats['uid']} bằng UID, {stats['name']} bằng tên.")

    # Fallback cuối: ghép theo thứ tự khi số lượng khớp 1-1
    used = {_s(r.get("_uid")) for r in all_data if _s(r.get("SĐT"))}
    pending_rows = [r for r in all_data if not _s(r.get("SĐT"))]
    pending_uids = [u for u in uid_to_phone if u not in used]

    if allow_positional and pending_rows and len(pending_rows) == len(pending_uids):
        if verbose:
            print(f"[CẢNH BÁO] Ghép {len(pending_rows)} dòng theo thứ tự "
                  f"(không có UID/tên để đối chiếu) - verify trước khi gọi.")
        for row, uid in zip(pending_rows, pending_uids):
            if apply_phone_info(row, uid_to_phone[uid], uid=uid,
                                real_name=uid_map.get(uid, "")):
                stats["positional"] += 1
    else:
        stats["pending"] = len(pending_rows)
        if verbose and pending_rows:
            print(f"-> Còn {len(pending_rows)} lead chưa có SĐT "
                  f"({len(pending_uids)} SĐT chưa dùng, số lượng lệch nên bỏ qua).")

    # Backfill tên nick thật cho mọi row còn là placeholder
    for row in all_data:
        uid = _s(row.get("_uid"))
        if uid and is_placeholder_name(row.get("Tên KH")) and uid_map.get(uid):
            row["Tên KH"] = uid_map[uid]

    return stats


# --------------------------------------------------------------------------- #
# MERGE / DEDUPE
# --------------------------------------------------------------------------- #

def lead_key(row: Dict[str, Any]) -> str:
    """Khoá gộp có tầng ưu tiên.

    Dòng CÓ uid và dòng KHÔNG uid nằm ở namespace khác nhau (prefix khác),
    nên không bao giờ bị gộp nhầm khi chưa đủ bằng chứng.
    """
    uid = _s(row.get("_uid")) or uid_from_url(row.get("Facebook"))
    if uid:
        return f"uid:{uid}"
    url = _s(row.get("Facebook"))
    if url:
        return f"url:{url.rstrip('/').lower()}"
    return f"name:{norm_name_key(row.get('Tên KH'))}"


def merge_rows(rows: Iterable[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Gộp nhiều record của cùng 1 khách thành 1 dòng.

    - Comment nhiều lần  -> nối bằng ' | ', không ghi đè
    - 'Số lần tìm'       -> số comment đã gộp (sale ưu tiên người hỏi nhiều lần)
    - Field khác         -> giữ giá trị non-empty đầu tiên
    - Tên nick thật      -> luôn thắng placeholder ('UID 100...', 'Người dùng Facebook')
    """
    merged: Dict[str, Dict[str, Any]] = {}
    order: List[str] = []

    def hits_of(r: Dict[str, Any]) -> int:
        """Số lần đã gộp trước đó — để merge_rows() gọi lại nhiều lần vẫn đúng
        (save_excel chạy lặp trong vòng lặp scrape)."""
        raw = _s(r.get("Số lần tìm"))
        return int(raw) if raw.isdigit() and int(raw) > 0 else 1

    for row in rows:
        key = lead_key(row)
        if key not in merged:
            new = dict(row)
            new["_comments"] = [c for c in [_s(row.get("Comment"))] if c]
            new["_hits"] = hits_of(row)
            merged[key] = new
            order.append(key)
            continue

        tgt = merged[key]
        tgt["_hits"] += hits_of(row)

        comment = _s(row.get("Comment"))
        if comment and comment not in tgt["_comments"]:
            tgt["_comments"].append(comment)

        # Tên thật thắng placeholder
        if is_placeholder_name(tgt.get("Tên KH")) and not is_placeholder_name(row.get("Tên KH")):
            tgt["Tên KH"] = _s(row.get("Tên KH"))

        for col in COLUMNS:
            if col == "Comment":
                continue
            if not _s(tgt.get(col)) and _s(row.get(col)):
                tgt[col] = _s(row.get(col))

    out: List[Dict[str, Any]] = []
    for key in order:
        r = merged[key]
        r["Comment"] = COMMENT_JOIN.join(r.pop("_comments"))
        r["Số lần tìm"] = r.pop("_hits")
        out.append(r)
    return out


# --------------------------------------------------------------------------- #
# EXCEL WRITER
# --------------------------------------------------------------------------- #

def _style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF", size=11)
    align = Alignment(horizontal="center", vertical="center")
    for idx, _ in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font, cell.fill, cell.alignment = font, fill, align
        ws.column_dimensions[get_column_letter(idx)].width = COL_WIDTHS[idx - 1]
    ws.row_dimensions[1].height = 22


def save_excel(all_data: Iterable[Dict[str, Any]], output_filename: str,
               merge: bool = True) -> bool:
    """Ghi Excel an toàn (atomic) — file cũ chỉ bị thay khi file mới ghi xong.

    merge=False dùng cho lần lưu giữa chừng nếu muốn giữ nguyên số dòng thô.
    """
    rows = list(all_data)
    if merge:
        rows = merge_rows(rows)

    tmp = f"{output_filename}.tmp"
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Danh sách tìm"

        ws.append(COLUMNS)
        _style_header(ws)

        wrap = Alignment(vertical="top", wrap_text=True)
        top = Alignment(vertical="top")
        link_font = Font(color="0563C1", underline="single")

        for r_idx, row in enumerate(rows, start=2):
            for c_idx, col in enumerate(COLUMNS, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                value = row.get(col, "")

                if col in PHONE_COLS:
                    # number_format phải set TRƯỚC khi gán, và value phải là str,
                    # nếu không Excel nuốt số 0 đầu.
                    cell.number_format = "@"
                    cell.value = norm_phone(value)
                elif col == "Số lần tìm":
                    cell.value = int(value) if str(value).isdigit() else ""
                elif col == "Facebook":
                    url = _s(value)
                    cell.value = url
                    if url.startswith("http"):
                        cell.hyperlink = url
                        cell.font = link_font
                elif col == "Zalo":
                    # Link nhảy thẳng sang Zalo kết bạn: https://zalo.me/<sđt>
                    phone = norm_phone(row.get("SĐT", ""))
                    if phone:
                        url = f"https://zalo.me/{phone}"
                        cell.value = url
                        cell.hyperlink = url
                        cell.font = link_font
                    else:
                        cell.value = ""
                elif col == "Link bài viết":
                    url = _s(value)
                    cell.value = url
                    if url.startswith("http"):
                        cell.hyperlink = url
                        cell.font = link_font
                else:
                    cell.value = _s(value)

                cell.alignment = wrap if col in WRAP_COLS else top

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(ws.max_row, 1)}"

        wb.save(tmp)
        os.replace(tmp, output_filename)   # atomic: file cũ còn nguyên nếu ghi lỗi
        print(f"[TIẾN ĐỘ] Đã lưu {len(rows)} khách hàng vào Excel.")
        return True

    except PermissionError:
        print("[CẢNH BÁO] Vui lòng đóng file Excel đang mở để lưu.")
    except Exception as e:
        print(f"[CẢNH BÁO] Lỗi ghi Excel: {e}")
    finally:
        if os.path.exists(tmp):
            try:
                os.remove(tmp)
            except OSError:
                pass
    return False
