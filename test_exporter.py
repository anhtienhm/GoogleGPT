#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
test_exporter.py — Test regression cho pipeline scrape -> Excel.

Chay khong can Selenium, khong can mang, khong can token that.
Dung sau moi lan sua app.py / exporter.py / run_hermes.py.

    python test_exporter.py        # exit 0 = pass, exit 1 = co test fail

Phu 8 nhom, moi nhom gan voi mot pitfall da tung xay ra (xem CLAUDE.md).
"""

import os
import sys
import tempfile
from pathlib import Path

# config.py doc token tu env; dat gia tri gia de import khong phu thuoc .env
os.environ.setdefault("FB_NUMBER_TOKEN", "dummy-token-for-tests")
sys.path.insert(0, str(Path(__file__).resolve().parent))

from bs4 import BeautifulSoup

import app
import exporter
from exporter import (
    COLUMNS,
    COL_WIDTHS,
    apply_phone_info,
    merge_phone_lookup,
    norm_phone,
    save_excel,
)

POST_URL = "https://www.facebook.com/groups/123/permalink/456/"

_failures = []


def check(name, cond, extra=""):
    print(("  OK   " if cond else "  FAIL ") + name + ("" if cond else f"  <- {extra}"))
    if not cond:
        _failures.append(name)


def article(html):
    return BeautifulSoup(html, "html.parser").find("div", attrs={"role": "article"})


# --------------------------------------------------------------------------- #
# 1. clean_facebook_url — pitfall #1 (regex bi nhan doi backslash -> luon N/A)
# --------------------------------------------------------------------------- #
print("\n[1] clean_facebook_url")
for url, want in [
    ("/profile.php?id=100012345678", "https://www.facebook.com/profile.php?id=100012345678"),
    ("https://www.facebook.com/profile.php?id=999&sk=about", "https://www.facebook.com/profile.php?id=999"),
    ("/groups/98765/user/100055512345/", "https://www.facebook.com/profile.php?id=100055512345"),
    ("/people/Bich-Quan-Nguyen/pfbid02E8MPFrJKNHftphRYAtkRS3xKfJ6QRhUEdEEGkrJthQX1VMx1CBjMRoL1qJPAUTxcl/",
     "https://www.facebook.com/profile.php?id=pfbid02E8MPFrJKNHftphRYAtkRS3xKfJ6QRhUEdEEGkrJthQX1VMx1CBjMRoL1qJPAUTxcl"),
    ("/nguyen.van.a", "https://www.facebook.com/nguyen.van.a"),
    ("https://m.facebook.com/tien.hoang.99", "https://www.facebook.com/tien.hoang.99"),
    ("/groups/98765/posts/111/", "N/A"),
    ("/watch", "N/A"),
]:
    got = app.clean_facebook_url(url)
    check(f"{url} -> {got}", got == want, f"mong doi {want}")

print("[1b] extract_uid_from_url / uid_from_url — UID pfbid (FB 2026)")
PFBID = "pfbid02E8MPFrJKNHftphRYAtkRS3xKfJ6QRhUEdEEGkrJthQX1VMx1CBjMRoL1qJPAUTxcl"
for url, want in [
    (f"https://www.facebook.com/people/Bich-Quan-Nguyen/{PFBID}/", PFBID),
    (f"https://www.facebook.com/profile.php?id={PFBID}", PFBID),
    ("https://www.facebook.com/people/Le-Van-C/100012345678/", "100012345678"),
    ("https://www.facebook.com/groups/98765/user/100055512345/", "100055512345"),
    ("https://www.facebook.com/nguyen.van.a", ""),
]:
    got = app.extract_uid_from_url(url) or ""
    check(f"extract_uid_from_url({url[:50]}...) -> {got[:20]}", got == want, f"mong doi {want[:20]}")
    got2 = exporter.uid_from_url(url)
    check(f"uid_from_url({url[:50]}...) -> {got2[:20]}", got2 == want, f"mong doi {want[:20]}")


# --------------------------------------------------------------------------- #
# 2. robust_parse_comment — pitfall #2 (nested.extract() xoa mat profile link)
# --------------------------------------------------------------------------- #
print("\n[2] robust_parse_comment — links_before truoc nested.extract()")
row = app.robust_parse_comment(article("""
<div role="article">
  <a href="/nguyen.van.a">Nguyen Van A</a>
  <div dir="auto">Gia bao nhieu v shop, ib minh nhe</div>
  <div role="article">
    <a href="/tran.thi.b">Tran Thi B</a>
    <div dir="auto">reply cua nguoi khac</div>
  </div>
</div>"""), POST_URL)

check("parse ra row", row is not None)
if row:
    check("Ten KH lay tu link cha", row["Tên KH"] == "Nguyen Van A", row["Tên KH"])
    check("link profile khong bi mat", row["Facebook"] == "https://www.facebook.com/nguyen.van.a", row["Facebook"])
    check("reply con bi loai khoi Comment", "reply cua nguoi khac" not in row["Comment"], row["Comment"])
    check("co key 'SĐT'", "SĐT" in row, list(row))
    check("co key 'Link bài viết'", row.get("Link bài viết") == POST_URL, row.get("Link bài viết"))


# --------------------------------------------------------------------------- #
# 3. Fallback ten trong group + trich SDT/UID tu comment
# --------------------------------------------------------------------------- #
print("\n[3] Fallback ten (group khong render <a>) + trich SDT/UID")
row = app.robust_parse_comment(article(
    '<div role="article"><a href="/profile.php?id=100099988877">Le Van C</a>'
    '<div dir="auto">Ib gia giup m, sdt 0776791717</div></div>'), POST_URL)
check("row parse duoc", row is not None)
if row:
    # extract_phone() tra ve thieu so 0 dau; norm_phone() va lai luc ghi Excel
    check("SĐT khong rong", bool(row["SĐT"]), row["SĐT"])
    check("SĐT sau norm_phone dung", norm_phone(row["SĐT"]) == "0776791717", norm_phone(row["SĐT"]))
    check("_uid trich duoc", row["_uid"] == "100099988877", row["_uid"])

row = app.robust_parse_comment(article(
    '<div role="article"><div dir="auto">Tran Thi Bich</div>'
    '<div dir="auto">ib gia bao nhieu shop</div></div>'), POST_URL)
check("lay duoc ten khi khong co the <a>", row is not None and row["Tên KH"] == "Tran Thi Bich",
      row["Tên KH"] if row else None)

# Regression (PR #1, Codex): article chi co DUY NHAT 1 doan dir=auto -> do la
# noi dung comment, KHONG phai ten. Lay nham lam ten thi vong boc noi dung se
# loai chinh no (text == author_name) -> comment rong -> lead bi bo IM LANG.
row = app.robust_parse_comment(article(
    '<div role="article"><div dir="auto">ib gia giup minh</div></div>'), POST_URL)
check("1 doan dir=auto duy nhat -> KHONG bi bo", row is not None, "row = None, lead bi bo")
if row:
    check("  giu comment lam noi dung", row["Comment"] == "ib gia giup minh", row["Comment"])
    check("  ten lui ve placeholder", row["Tên KH"] == "Nguoi dung Facebook", row["Tên KH"])

# dir=auto long nhau, ben trong cung chi la comment -> van khong duoc lay lam ten
row = app.robust_parse_comment(article(
    '<div role="article"><div dir="auto"><span dir="auto">ib gia</span></div></div>'), POST_URL)
check("dir=auto long nhau, chi co comment -> KHONG bi bo",
      row is not None and row["Tên KH"] == "Nguoi dung Facebook",
      "None" if row is None else row["Tên KH"])

# Regression (PR #1, Codex lan 2): Facebook tach 1 comment thanh NHIEU block
# dir=auto. cands[0] la dong dau cua comment, khong phai ten. Lay lam ten thi
# tu khoa mua hang nam trong do bi mat -> lead bi bo, hoac Ten KH la manh comment.
row = app.robust_parse_comment(article(
    '<div role="article"><div dir="auto">ib gia bao nhieu</div>'
    '<div dir="auto">ship ve Ha Noi duoc khong</div></div>'), POST_URL)
check("comment tach 2 block, tu khoa o block dau -> KHONG bi bo", row is not None,
      "row = None, lead bi bo")
if row:
    check("  ten lui ve placeholder", row["Tên KH"] == "Nguoi dung Facebook", row["Tên KH"])
    check("  giu ca 2 block lam comment",
          "ib gia bao nhieu" in row["Comment"] and "ship ve Ha Noi" in row["Comment"], row["Comment"])

# Ca 2 block deu co tu khoa -> van khong duoc lay block dau lam ten
row = app.robust_parse_comment(article(
    '<div role="article"><div dir="auto">ib gia</div>'
    '<div dir="auto">tu van giup minh</div></div>'), POST_URL)
check("2 block deu co tu khoa -> Ten KH KHONG phai manh comment",
      row is not None and row["Tên KH"] == "Nguoi dung Facebook",
      "None" if row is None else row["Tên KH"])


# --------------------------------------------------------------------------- #
# 4. TIME_PATTERN phai khop timestamp CO DAU (Facebook render co dau)
# --------------------------------------------------------------------------- #
print("\n[4] TIME_PATTERN — dau tieng Viet")
for t in ["2 giờ", "15 phút", "3 ngày", "1 tuần", "Vừa xong"]:
    check(f"khop {t!r}", bool(app.TIME_PATTERN.match(t)))
check("is_junk('2 giờ') loc duoc timestamp", app.is_junk("2 giờ"))


# --------------------------------------------------------------------------- #
# 5. apply_phone_info — pitfall #3 (KHONG ghi de SDT lay tu comment)
# --------------------------------------------------------------------------- #
print("\n[5] apply_phone_info — khong ghi de SĐT da co")
info = {"number": "0900000000", "number2": "0911111111", "numberProvider": "Viettel",
        "number2Provider": "Mobi", "location": "Ha Noi", "gender": "Nam", "birthday": "01/01/1990"}

r = {"SĐT": "0776791717", "Tên KH": "Le Van C"}
apply_phone_info(r, info)
check("SĐT goc giu nguyen", r["SĐT"] == "0776791717", r["SĐT"])
check("SĐT 2 duoc bo sung", r.get("SĐT 2") == "0911111111", r.get("SĐT 2"))
check("Location duoc bo sung", r.get("Location") == "Ha Noi", r.get("Location"))

r = {"SĐT": "", "Tên KH": "X"}
apply_phone_info(r, info)
check("row trong -> nhan SĐT tu API", r["SĐT"] == "0900000000", r["SĐT"])


# --------------------------------------------------------------------------- #
# 6. merge_phone_lookup — KHONG skip row da co SDT
# --------------------------------------------------------------------------- #
print("\n[6] merge_phone_lookup — row co SĐT van duoc lam giau")
all_data = [
    {"Tên KH": "Le Van C", "SĐT": "0776791717", "_uid": "100099988877",
     "Comment": "ib gia", "Link bài viết": POST_URL},
    {"Tên KH": "Nguyen Van A", "SĐT": "", "_uid": "100012345678",
     "Comment": "bao nhieu", "Link bài viết": POST_URL},
]
merge_phone_lookup(
    all_data,
    {"100099988877": "Le Van C", "100012345678": "Nguyen Van A"},
    {"100099988877": info,
     "100012345678": {"number": "0988888888", "number2": "", "numberProvider": "Vina",
                      "number2Provider": "", "location": "HCM", "gender": "Nu", "birthday": ""}},
)
check("row co SĐT van nhan SĐT 2", all_data[0].get("SĐT 2") == "0911111111", all_data[0].get("SĐT 2"))
check("row co SĐT giu so goc", all_data[0]["SĐT"] == "0776791717", all_data[0]["SĐT"])
check("row trong nhan SĐT tu API", all_data[1]["SĐT"] == "0988888888", all_data[1]["SĐT"])


# --------------------------------------------------------------------------- #
# 7. Excel — pitfall #4 (key phai khop COLUMNS) va so 0 dau cua SDT
# --------------------------------------------------------------------------- #
print("\n[7] save_excel — 15 cot, hyperlink, SĐT giu so 0")
check("'Link bài viết' co trong COLUMNS", "Link bài viết" in COLUMNS)
check("COLUMNS = 15 cot", len(COLUMNS) == 15, len(COLUMNS))
check("COL_WIDTHS khop COLUMNS", len(COL_WIDTHS) == len(COLUMNS), f"{len(COL_WIDTHS)} vs {len(COLUMNS)}")

import openpyxl

with tempfile.TemporaryDirectory() as tmp:
    out = str(Path(tmp) / "out.xlsx")
    save_excel(all_data, out)
    ws = openpyxl.load_workbook(out).active
    check("header Excel = COLUMNS", [c.value for c in ws[1]] == COLUMNS)

    cell = ws.cell(row=2, column=COLUMNS.index("Link bài viết") + 1)
    check("cot Link bài viết co gia tri", cell.value == POST_URL, cell.value)
    check("cot Link bài viết la hyperlink", cell.hyperlink is not None)

    sdt = ws.cell(row=2, column=COLUMNS.index("SĐT") + 1)
    check("SĐT la text, giu so 0 dau", sdt.value == "0776791717" and isinstance(sdt.value, str), sdt.value)


# --------------------------------------------------------------------------- #
# 8. Hop dong cot voi run_hermes.py + doc lai bang pandas
# --------------------------------------------------------------------------- #
print("\n[8] Hop dong cot run_hermes.py")
try:
    import pandas as pd
    import run_hermes as H

    with tempfile.TemporaryDirectory() as tmp:
        out = str(Path(tmp) / "h.xlsx")
        save_excel(all_data, out)
        df = pd.read_excel(out, dtype=str)      # dtype=str: khong de pandas ep SDT thanh float

        for label, col in [("COL_NAME", H.COL_NAME), ("COL_COMMENT", H.COL_COMMENT),
                           ("COL_PHONE", H.COL_PHONE), ("COL_USER_LINK", H.COL_USER_LINK),
                           ("COL_POST_LINK", H.COL_POST_LINK), ("COL_TIME", H.COL_TIME)]:
            check(f"{label} = {col!r} co trong Excel", col in df.columns,
                  "khong co -> Hermes doc ra N/A")

        # pandas mac dinh ep "0776791717" -> 776791717.0 (mat so 0, them ".0")
        raw = pd.read_excel(out)[H.COL_PHONE].iloc[0]
        check("norm_phone va duoc so bi pandas ep thanh float",
              norm_phone(str(raw).replace(".0", "")) == "0776791717", raw)
except ImportError as exc:
    print(f"  BO QUA — thieu pandas ({exc})")


# --------------------------------------------------------------------------- #
print("\n" + "=" * 58)
if _failures:
    print(f"CO {len(_failures)} TEST FAIL:")
    for f in _failures:
        print("  -", f)
    sys.exit(1)
print("TAT CA TEST PASS")
sys.exit(0)
